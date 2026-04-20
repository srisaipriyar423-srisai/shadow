from __future__ import annotations

import json
import math
import os
from collections import Counter
from collections import defaultdict
from datetime import datetime
from io import BytesIO
from pathlib import Path
from statistics import mean

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.graphics.shapes import Circle, Drawing, Line, Rect, String
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from utils.helpers import format_duration
from utils.theme import THEME, apply_theme

BASE_DIR = Path(__file__).resolve().parents[1]
DATA_DIR = BASE_DIR / "data"
SESSIONS_PATH = DATA_DIR / "sessions.json"
ALERTS_PATH = DATA_DIR / "alerts.json"
SETTINGS_PATH = DATA_DIR / "settings.json"
CATALOG_PATH = DATA_DIR / "saas_catalog.json"

DARK_BG = "0C0A04"
PANEL_BG = "13100A"
GOLD = "FFD700"
CREAM = "E8D48A"
RED = "F44336"
AMBER = "FF9800"
GREEN = "4CAF50"


def parse_dt(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value
    text = str(value).replace("Z", "")
    fmts = ["%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"]
    for fmt in fmts:
        try:
            return datetime.strptime(text[:19], fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text)
    except ValueError:
        return None


def dt_text(value):
    parsed = parse_dt(value)
    return parsed.strftime("%Y-%m-%d %H:%M:%S") if parsed else "N/A"


def pages_text(pages):
    if not pages:
        return "N/A"
    return ", ".join(str(p) for p in pages)


def risk_rank(level):
    mapping = {"CLEAN": 0, "LOW": 1, "MODERATE": 2, "MEDIUM": 2, "HIGH": 3, "CRITICAL": 4}
    return mapping.get(str(level or "").upper(), 0)


def score_to_verdict(score):
    if score >= 85:
        return "CRITICAL", RED
    if score >= 66:
        return "HIGH", RED
    if score >= 31:
        return "MODERATE", AMBER
    if score >= 1:
        return "LOW", GREEN
    return "CLEAN", GREEN


def safe_json(path, default, section_name, warnings):
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        warnings.append(f"{section_name}: file not found ({path.name}).")
        return default
    except json.JSONDecodeError as ex:
        warnings.append(f"{section_name}: malformed JSON skipped ({path.name}) - {ex}")
        return default
    except Exception as ex:  # noqa: BLE001
        warnings.append(f"{section_name}: read error ({path.name}) - {ex}")
        return default


def normalize_tool(d):
    duration = d.get("duration") or {}
    return {
        "tool_name": d.get("tool_name", d.get("name", "N/A")),
        "category": d.get("category", "N/A"),
        "risk": str(d.get("risk_level", "LOW")).upper(),
        "query_count": int(d.get("query_count", 0) or 0),
        "unique_ips": int(d.get("unique_ips", 0) or 0),
        "gdpr": bool(d.get("gdpr_concern", False)),
        "risk_reasons": d.get("risk_reasons") or [],
        "alternative": d.get("approved_alternative", "N/A"),
        "domains": d.get("domains_matched") or d.get("domains") or [],
        "duration": {
            "first_query_at": duration.get("first_query_at", d.get("first_seen")),
            "last_query_at": duration.get("last_query_at", d.get("last_seen")),
            "total_span_seconds": int(duration.get("total_span_seconds", 0) or 0),
            "total_span_fmt": duration.get("total_span_fmt", "00h 00m 00s"),
            "active_minutes": int(duration.get("active_minutes", 0) or 0),
            "active_minutes_fmt": duration.get("active_minutes_fmt", "0h 00m"),
            "queries_per_minute": float(duration.get("queries_per_minute", 0) or 0),
            "peak_usage_hour": duration.get("peak_usage_hour"),
        },
    }


def extract_ip_rows(session):
    rows = []
    ip_activity = session.get("ip_activity") or session.get("source_ip_activity") or []
    for item in ip_activity:
        if not isinstance(item, dict):
            continue
        rows.append(
            {
                "ip": str(item.get("ip", "N/A")),
                "query_count": int(item.get("query_count", 0) or 0),
                "tools_accessed": item.get("tools_accessed") or item.get("tools") or [],
                "risk_level": str(item.get("risk_level", "N/A")).upper(),
            }
        )
    raw_ips = session.get("source_ips") or []
    for ip in raw_ips:
        rows.append({"ip": str(ip), "query_count": 0, "tools_accessed": [], "risk_level": "N/A"})
    unique = {}
    for row in rows:
        ip = row["ip"]
        if ip not in unique:
            unique[ip] = row
        else:
            unique[ip]["query_count"] += row["query_count"]
            unique[ip]["tools_accessed"] = sorted(set(unique[ip]["tools_accessed"] + row["tools_accessed"]))
            if risk_rank(row["risk_level"]) > risk_rank(unique[ip]["risk_level"]):
                unique[ip]["risk_level"] = row["risk_level"]
    return list(unique.values())


def normalize_session(session, index):
    tools = [normalize_tool(t) for t in (session.get("tool_detections") or [])]
    ip_rows = extract_ip_rows(session)
    parsed_at = session.get("uploaded_at", session.get("parsed_at", "N/A"))
    total_queries = int(session.get("log_lines", session.get("total_dns_queries", session.get("query_count", 0))) or 0)
    unique_domains = int(session.get("unique_domains", 0) or 0)
    unique_source_ips = int(session.get("unique_source_ips", session.get("source_ip_count", 0)) or 0)
    if not unique_source_ips and ip_rows:
        unique_source_ips = len({r["ip"] for r in ip_rows})

    high_count = sum(1 for t in tools if t["risk"] == "HIGH")
    med_count = sum(1 for t in tools if t["risk"] == "MEDIUM")
    low_count = sum(1 for t in tools if t["risk"] == "LOW")
    gdpr_count = sum(1 for t in tools if t["gdpr"])

    score = float(session.get("exposure_score", 0) or 0)
    verdict, _ = score_to_verdict(score)

    user_session = session.get("user_session") or {}
    page_visit_log = user_session.get("page_visit_log") or []
    pages_visited = user_session.get("pages_visited") or []
    if not pages_visited:
        pages_visited = [
            str(item.get("page", "")).strip()
            for item in page_visit_log
            if isinstance(item, dict) and str(item.get("page", "")).strip()
        ]
        pages_visited = list(dict.fromkeys(pages_visited))

    page_counts = Counter(
        str(item.get("page", "")).strip()
        for item in page_visit_log
        if isinstance(item, dict) and str(item.get("page", "")).strip()
    )
    most_used_page = page_counts.most_common(1)[0][0] if page_counts else "N/A"

    analysis_duration_sec = int(
        user_session.get(
            "analysis_duration_sec",
            session.get("analysis_duration_seconds", 0),
        )
        or 0
    )
    total_app_duration_sec = int(user_session.get("total_app_duration_sec", 0) or 0)
    if total_app_duration_sec == 0:
        app_opened_dt = parse_dt(user_session.get("app_opened_at"))
        app_closed_dt = parse_dt(user_session.get("app_closed_at"))
        if app_opened_dt and app_closed_dt:
            total_app_duration_sec = int((app_closed_dt - app_opened_dt).total_seconds())

    top_domains = []
    for entry in (session.get("unknown_domains") or [])[:3]:
        if isinstance(entry, dict):
            top_domains.append((entry.get("domain", "N/A"), int(entry.get("query_count", 0) or 0)))

    return {
        "session_id": session.get("id", f"session_{index}"),
        "name": session.get("name", f"Session {index}"),
        "parsed_at": parsed_at,
        "source_file": session.get("source_file", session.get("file_name", session.get("name", "N/A"))),
        "file_size": session.get("file_size", session.get("file_size_bytes", "N/A")),
        "total_queries": total_queries,
        "unique_domains": unique_domains,
        "unique_source_ips": unique_source_ips,
        "shadow_tools": len(tools),
        "high_count": high_count,
        "medium_count": med_count,
        "low_count": low_count,
        "gdpr_count": gdpr_count,
        "score": score,
        "score_label": verdict,
        "auto_alerts": int(session.get("auto_alerts_created", 0) or 0),
        "app_opened_at": user_session.get("app_opened_at"),
        "analysis_started_at": user_session.get("analysis_started_at", session.get("analysis_start_time")),
        "analysis_ended_at": user_session.get("analysis_ended_at", session.get("analysis_end_time")),
        "app_closed_at": user_session.get("app_closed_at"),
        "analysis_duration_sec": analysis_duration_sec,
        "analysis_duration_fmt": user_session.get("analysis_duration_fmt", format_duration(analysis_duration_sec)),
        "total_app_duration_sec": total_app_duration_sec,
        "total_app_duration_fmt": user_session.get("total_app_duration_fmt", format_duration(total_app_duration_sec)),
        "pages_visited": pages_visited,
        "page_visit_log": page_visit_log,
        "page_visit_count": len(page_visit_log),
        "most_used_page": most_used_page,
        "ip_tool_durations": session.get("ip_tool_durations", []),
        "tool_hourly_usage": session.get("tool_hourly_usage", []),
        "ip_hourly_usage": session.get("ip_hourly_usage", []),
        "tools": tools,
        "top_domains": top_domains,
        "ip_rows": ip_rows,
    }


def compute_report(selected_sessions, all_sessions, alerts, settings, catalog_tools, org_name, report_title):
    normalized = [normalize_session(s, idx + 1) for idx, s in enumerate(selected_sessions)]
    all_normalized = [normalize_session(s, idx + 1) for idx, s in enumerate(all_sessions)]

    dates = [parse_dt(s["parsed_at"]) for s in normalized if parse_dt(s["parsed_at"]) is not None]
    earliest = min(dates).strftime("%Y-%m-%d %H:%M:%S") if dates else "N/A"
    latest = max(dates).strftime("%Y-%m-%d %H:%M:%S") if dates else "N/A"

    unique_domains = set()
    unique_ips = set()
    tool_names = set()
    score_list = []
    total_queries = 0
    medium_threshold_exceeded = 0
    total_shadow_app_usage_sec = 0
    global_hourly_queries = Counter()
    after_hours_tool_names = set()
    longest_single_tool_session = None
    total_app_duration_sec = 0
    analysis_duration_sec_total = 0
    session_durations = []
    session_start_hours = Counter()
    overall_page_counter = Counter()

    for s in normalized:
        total_queries += s["total_queries"]
        score_list.append(s["score"])
        if s["score"] >= 31:
            medium_threshold_exceeded += 1
        total_app_duration_sec += int(s.get("total_app_duration_sec", 0) or 0)
        analysis_duration_sec_total += int(s.get("analysis_duration_sec", 0) or 0)
        if int(s.get("total_app_duration_sec", 0) or 0) > 0:
            session_durations.append(s)
        app_opened_dt = parse_dt(s.get("app_opened_at"))
        if app_opened_dt:
            session_start_hours[app_opened_dt.hour] += 1
        for page_name in s.get("pages_visited", []):
            overall_page_counter[str(page_name)] += 1
        for visit in s.get("page_visit_log", []):
            if isinstance(visit, dict) and visit.get("page"):
                overall_page_counter[str(visit.get("page"))] += 1
        for t in s["tools"]:
            tool_names.add(t["tool_name"])
            tool_duration = t.get("duration", {})
            d_secs = int(tool_duration.get("total_span_seconds", 0) or 0)
            total_shadow_app_usage_sec += d_secs
            peak_hr = tool_duration.get("peak_usage_hour")
            if isinstance(peak_hr, int):
                global_hourly_queries[peak_hr] += int(t.get("query_count", 0) or 0)
                if peak_hr < 9 or peak_hr >= 18:
                    after_hours_tool_names.add(t["tool_name"])
            if d_secs > 0:
                cand = {
                    "tool_name": t["tool_name"],
                    "duration_sec": d_secs,
                    "duration_fmt": tool_duration.get("total_span_fmt", format_duration(d_secs)),
                    "session_parsed_at": s.get("parsed_at"),
                }
                if not longest_single_tool_session or cand["duration_sec"] > longest_single_tool_session["duration_sec"]:
                    longest_single_tool_session = cand
            for d in t["domains"]:
                unique_domains.add(str(d))
        for ip in s["ip_rows"]:
            if ip["ip"] != "N/A":
                unique_ips.add(ip["ip"])
        for dom, _count in s["top_domains"]:
            unique_domains.add(dom)

    overall_score = round(mean(score_list), 2) if score_list else 0.0
    overall_verdict, verdict_color = score_to_verdict(overall_score)

    avg_session_duration_sec = int(total_app_duration_sec / len(normalized)) if normalized else 0
    longest_session = max(session_durations, key=lambda x: x.get("total_app_duration_sec", 0), default=None)
    shortest_session = min(session_durations, key=lambda x: x.get("total_app_duration_sec", 0), default=None)
    most_visited_page = overall_page_counter.most_common(1)[0][0] if overall_page_counter else "N/A"
    peak_hour = "N/A"
    if session_start_hours:
        peak_hr = session_start_hours.most_common(1)[0][0]
        peak_hour = f"{peak_hr:02d}:00-{(peak_hr + 1) % 24:02d}:00"
    shadow_peak_hour = peak_hour
    if global_hourly_queries:
        peak_shadow_hr = global_hourly_queries.most_common(1)[0][0]
        shadow_peak_hour = f"{peak_shadow_hr:02d}:00-{(peak_shadow_hr + 1) % 24:02d}:00"

    session_timing_rows = []
    for idx, s in enumerate(normalized, start=1):
        session_timing_rows.append(
            {
                "Session #": idx,
                "Session ID": s["session_id"],
                "Session Date": dt_text(s.get("parsed_at")),
                "App Opened At": dt_text(s.get("app_opened_at")),
                "Analysis Start": dt_text(s.get("analysis_started_at")),
                "Analysis End": dt_text(s.get("analysis_ended_at")),
                "App Closed At": dt_text(s.get("app_closed_at")),
                "Analysis Duration": s.get("analysis_duration_fmt", "N/A"),
                "Analysis Duration Sec": int(s.get("analysis_duration_sec", 0) or 0),
                "Total App Duration": s.get("total_app_duration_fmt", "N/A"),
                "Total App Duration Sec": int(s.get("total_app_duration_sec", 0) or 0),
                "Pages Visited": pages_text(s.get("pages_visited", [])),
                "Page Visit Count": int(s.get("page_visit_count", 0) or 0),
                "Most Used Page": s.get("most_used_page", "N/A"),
            }
        )

    ordered_by_date = sorted(normalized, key=lambda x: parse_dt(x["parsed_at"]) or datetime.min)
    highest = max(ordered_by_date, key=lambda x: x["score"], default=None)
    lowest = min(ordered_by_date, key=lambda x: x["score"], default=None)

    trend_symbol = "→"
    trend_text = "Stable"
    if len(ordered_by_date) >= 3:
        last_three = [s["score"] for s in ordered_by_date[-3:]]
        if last_three[-1] - last_three[0] > 5:
            trend_symbol = "↑"
            trend_text = "Rising"
        elif last_three[-1] - last_three[0] < -5:
            trend_symbol = "↓"
            trend_text = "Falling"

    first_high_seen = set()
    high_new_count = 0
    for s in ordered_by_date:
        for t in s["tools"]:
            if t["risk"] == "HIGH" and t["tool_name"] not in first_high_seen:
                high_new_count += 1
                first_high_seen.add(t["tool_name"])

    device_map = {}
    for s in normalized:
        when = parse_dt(s["parsed_at"])
        for ip in s["ip_rows"]:
            ip_addr = ip["ip"]
            if ip_addr == "N/A":
                continue
            rec = device_map.setdefault(
                ip_addr,
                {
                    "ip": ip_addr,
                    "queries": 0,
                    "sessions": set(),
                    "tools": set(),
                    "highest_risk": "LOW",
                    "gdpr_tools": set(),
                    "first_seen": when,
                    "last_seen": when,
                    "high_session_hits": set(),
                },
            )
            rec["queries"] += ip["query_count"]
            rec["sessions"].add(s["session_id"])
            for tool in ip["tools_accessed"]:
                rec["tools"].add(str(tool))
            if risk_rank(ip["risk_level"]) > risk_rank(rec["highest_risk"]):
                rec["highest_risk"] = ip["risk_level"]
            if when:
                rec["first_seen"] = min(rec["first_seen"], when) if rec["first_seen"] else when
                rec["last_seen"] = max(rec["last_seen"], when) if rec["last_seen"] else when

        high_tools = [t for t in s["tools"] if t["risk"] == "HIGH"]
        gdpr_tools = [t for t in s["tools"] if t["gdpr"]]
        for ip in s["ip_rows"]:
            rec = device_map.get(ip["ip"])
            if not rec:
                continue
            for t in gdpr_tools:
                rec["gdpr_tools"].add(t["tool_name"])
            if high_tools:
                rec["high_session_hits"].add(s["session_id"])

    device_rows = []
    for rec in device_map.values():
        device_rows.append(
            {
                "IP Address": rec["ip"],
                "Total Queries (all sessions)": rec["queries"],
                "Sessions Active": len(rec["sessions"]),
                "Shadow Tools Used": len(rec["tools"]),
                "Highest Risk Level": rec["highest_risk"],
                "GDPR Tools Accessed": ", ".join(sorted(rec["gdpr_tools"])) or "N/A",
                "First Seen": rec["first_seen"].strftime("%Y-%m-%d %H:%M:%S") if rec["first_seen"] else "N/A",
                "Last Seen": rec["last_seen"].strftime("%Y-%m-%d %H:%M:%S") if rec["last_seen"] else "N/A",
                "repeat_offender": len(rec["high_session_hits"]) >= 2,
            }
        )
    device_rows.sort(key=lambda x: x["Total Queries (all sessions)"], reverse=True)

    power_users = device_rows[:10]
    gdpr_devices = [r for r in device_rows if r["GDPR Tools Accessed"] != "N/A"]
    repeat_offenders = [r for r in device_rows if r.get("repeat_offender")]

    per_device_app_rows = []
    ip_hour_heatmap_counter: dict[str, Counter] = defaultdict(Counter)
    tool_hour_heatmap_counter: dict[str, Counter] = defaultdict(Counter)
    for s in normalized:
        session_label = s.get("parsed_at", "N/A")
        for row in s.get("ip_tool_durations", []):
            per_device_app_rows.append(
                {
                    "Session": session_label,
                    "IP Address": row.get("ip", "N/A"),
                    "Tool Name": row.get("tool", "N/A"),
                    "Category": row.get("category", "N/A"),
                    "Risk": row.get("risk", "N/A"),
                    "First Seen": dt_text(row.get("first_query_at")),
                    "Last Seen": dt_text(row.get("last_query_at")),
                    "Total Duration": row.get("total_span_fmt", "00h 00m 00s"),
                    "Total Duration Sec": int(row.get("total_span_seconds", 0) or 0),
                    "Active Minutes": int(row.get("active_minutes", 0) or 0),
                    "Query Count": int(row.get("query_count", 0) or 0),
                }
            )
        for ip_row in s.get("ip_hourly_usage", []):
            ip_val = str(ip_row.get("ip", "N/A"))
            for h in range(24):
                ip_hour_heatmap_counter[ip_val][h] += int(ip_row.get(str(h), 0) or 0)
        for t_row in s.get("tool_hourly_usage", []):
            tool_name = str(t_row.get("tool_name", "N/A"))
            h_map = t_row.get("hourly_queries", {})
            for h in range(24):
                tool_hour_heatmap_counter[tool_name][h] += int(h_map.get(str(h), 0) or 0)

    per_device_app_rows.sort(key=lambda x: x.get("Total Duration Sec", 0), reverse=True)

    ip_heatmap_rows = []
    max_ip_heat = 0
    for ip, h_counter in ip_hour_heatmap_counter.items():
        row = {"IP Address": ip}
        total = 0
        for h in range(24):
            val = int(h_counter.get(h, 0))
            row[f"Hour {h:02d}"] = val
            total += val
            max_ip_heat = max(max_ip_heat, val)
        row["Total"] = total
        ip_heatmap_rows.append(row)
    for row in ip_heatmap_rows:
        row["Highest Cell Flag"] = "YES" if any(row.get(f"Hour {h:02d}", 0) == max_ip_heat and max_ip_heat > 0 for h in range(24)) else ""

    tool_heatmap_rows = []
    for tool, h_counter in tool_hour_heatmap_counter.items():
        row = {"Tool": tool}
        total = 0
        for h in range(24):
            val = int(h_counter.get(h, 0))
            row[f"Hour {h:02d}"] = val
            total += val
        row["Total"] = total
        tool_heatmap_rows.append(row)

    tool_map = {}
    sessions_sorted = sorted(normalized, key=lambda x: parse_dt(x["parsed_at"]) or datetime.min)
    for s in sessions_sorted:
        sdt = parse_dt(s["parsed_at"])
        for t in s["tools"]:
            key = t["tool_name"]
            rec = tool_map.setdefault(
                key,
                {
                    "Tool Name": key,
                    "Category": t["category"],
                    "Risk Level": t["risk"],
                    "Total Queries (all sessions)": 0,
                    "Sessions Appeared In": set(),
                    "Unique IPs": 0,
                    "GDPR Concern": "YES" if t["gdpr"] else "No",
                    "Risk Reasons (full list)": set(),
                    "Recommended Alternative": t["alternative"],
                    "First Detected": sdt,
                    "Last Detected": sdt,
                    "series": [],
                    "Total Duration Sec": 0,
                    "Longest Session Duration Sec": 0,
                    "Total Active Minutes": 0,
                    "Queries Per Minute Sum": 0.0,
                    "Peak Hour Counter": Counter(),
                    "After Hours Queries": 0,
                    "Total Queries For Hours": 0,
                    "Duration Timeline": [],
                },
            )
            rec["Total Queries (all sessions)"] += t["query_count"]
            rec["Sessions Appeared In"].add(s["session_id"])
            rec["Unique IPs"] += t["unique_ips"]
            rec["Risk Reasons (full list)"].update(t["risk_reasons"])
            rec["series"].append((s["session_id"], t["query_count"]))
            td = t.get("duration", {})
            d_sec = int(td.get("total_span_seconds", 0) or 0)
            rec["Total Duration Sec"] += d_sec
            rec["Longest Session Duration Sec"] = max(rec["Longest Session Duration Sec"], d_sec)
            rec["Total Active Minutes"] += int(td.get("active_minutes", 0) or 0)
            rec["Queries Per Minute Sum"] += float(td.get("queries_per_minute", 0) or 0)
            ph = td.get("peak_usage_hour")
            if isinstance(ph, int):
                rec["Peak Hour Counter"][ph] += int(t["query_count"])
            if isinstance(ph, int) and (ph < 9 or ph >= 18):
                rec["After Hours Queries"] += int(t["query_count"])
            rec["Total Queries For Hours"] += int(t["query_count"])
            rec["Duration Timeline"].append((s.get("parsed_at", "N/A"), d_sec))
            if t["gdpr"]:
                rec["GDPR Concern"] = "YES"
            if sdt:
                rec["First Detected"] = min(rec["First Detected"], sdt) if rec["First Detected"] else sdt
                rec["Last Detected"] = max(rec["Last Detected"], sdt) if rec["Last Detected"] else sdt
            if risk_rank(t["risk"]) > risk_rank(rec["Risk Level"]):
                rec["Risk Level"] = t["risk"]

    for rec in tool_map.values():
        if rec["Peak Hour Counter"]:
            rec["Peak Usage Hour"] = rec["Peak Hour Counter"].most_common(1)[0][0]
        else:
            rec["Peak Usage Hour"] = None
        total_q_for_hours = max(1, int(rec.get("Total Queries For Hours", 0) or 0))
        rec["After Hours Pct"] = (float(rec.get("After Hours Queries", 0) or 0) / total_q_for_hours) * 100

    master_tools = []
    for rec in tool_map.values():
        total_duration_sec = int(rec.get("Total Duration Sec", 0) or 0)
        sessions_appeared = len(rec["Sessions Appeared In"])
        avg_duration_sec = int(total_duration_sec / sessions_appeared) if sessions_appeared else 0
        longest_session_sec = int(rec.get("Longest Session Duration Sec", 0) or 0)
        master_tools.append(
            {
                "Tool Name": rec["Tool Name"],
                "Category": rec["Category"],
                "Risk Level": rec["Risk Level"],
                "Total Queries (all sessions)": rec["Total Queries (all sessions)"],
                "Sessions Appeared In": len(rec["Sessions Appeared In"]),
                "Unique IPs": rec["Unique IPs"],
                "GDPR Concern": rec["GDPR Concern"],
                "Risk Reasons (full list)": "; ".join(sorted(rec["Risk Reasons (full list)"])) or "N/A",
                "Recommended Alternative": rec["Recommended Alternative"],
                "First Detected": rec["First Detected"].strftime("%Y-%m-%d %H:%M:%S") if rec["First Detected"] else "N/A",
                "Last Detected": rec["Last Detected"].strftime("%Y-%m-%d %H:%M:%S") if rec["Last Detected"] else "N/A",
                "Total Duration Sec": total_duration_sec,
                "Total Duration (all sessions)": format_duration(total_duration_sec),
                "Avg Duration Per Session": format_duration(avg_duration_sec),
                "Longest Session Duration Sec": longest_session_sec,
                "Longest Single Session Duration": format_duration(longest_session_sec),
                "Total Active Minutes": int(rec.get("Total Active Minutes", 0) or 0),
                "Avg Queries/Min": round(float(rec.get("Queries Per Minute Sum", 0) or 0) / max(1, sessions_appeared), 2),
                "Peak Usage Hour": (
                    f"{int(rec.get('Peak Usage Hour', 0)):02d}:00-{(int(rec.get('Peak Usage Hour', 0)) + 1) % 24:02d}:00"
                    if rec.get("Peak Usage Hour") is not None
                    else "N/A"
                ),
                "After Hours %": round(float(rec.get("After Hours Pct", 0.0) or 0.0), 2),
                "After Hours Flag": "🌙" if float(rec.get("After Hours Pct", 0.0) or 0.0) > 20 else "",
                "Duration Timeline": rec.get("Duration Timeline", []),
                "_series": rec["series"],
            }
        )
    master_tools.sort(key=lambda x: x["Total Queries (all sessions)"], reverse=True)

    duration_master_rows = sorted(master_tools, key=lambda x: x.get("Total Duration Sec", 0), reverse=True)
    top_duration_tools = duration_master_rows[:5]
    intensive_tools = sorted(master_tools, key=lambda x: x.get("Avg Queries/Min", 0), reverse=True)[:5]
    after_hours_tools = sorted(master_tools, key=lambda x: x.get("After Hours %", 0), reverse=True)[:5]
    duration_timeline = [
        {
            "tool": t["Tool Name"],
            "points": t.get("Duration Timeline", []),
        }
        for t in master_tools
    ]

    longest_total_tool = top_duration_tools[0] if top_duration_tools else None
    longest_single_tool = max(master_tools, key=lambda x: x.get("Longest Session Duration Sec", 0), default=None)
    avg_shadow_usage_per_session_sec = int(total_shadow_app_usage_sec / len(normalized)) if normalized else 0
    pct_workday_shadow = round((avg_shadow_usage_per_session_sec / (8 * 3600)) * 100, 2) if normalized else 0.0

    most_persistent = max(master_tools, key=lambda x: x["Sessions Appeared In"], default=None)
    fastest_growing = None
    fastest_growth_delta = -10**9
    for t in master_tools:
        counts = [p[1] for p in t.get("_series", [])]
        for i in range(1, len(counts)):
            delta = counts[i] - counts[i - 1]
            if delta > fastest_growth_delta:
                fastest_growth_delta = delta
                fastest_growing = t

    latest_session = sessions_sorted[-1] if sessions_sorted else None
    latest_tool_names = {t["tool_name"] for t in (latest_session["tools"] if latest_session else [])}
    previous_tool_names = set()
    for s in sessions_sorted[:-1]:
        previous_tool_names.update({t["tool_name"] for t in s["tools"]})
    new_tools_latest = sorted(latest_tool_names - previous_tool_names)
    disappeared_tools = sorted(previous_tool_names - latest_tool_names)

    category_map = defaultdict(lambda: {"tools": set(), "queries": 0, "highest": "LOW", "gdpr_count": 0})
    total_shadow_queries = sum(x["Total Queries (all sessions)"] for x in master_tools) or 1
    for tool in master_tools:
        cat = tool["Category"]
        rec = category_map[cat]
        rec["tools"].add(tool["Tool Name"])
        rec["queries"] += tool["Total Queries (all sessions)"]
        rec["gdpr_count"] += 1 if tool["GDPR Concern"] == "YES" else 0
        if risk_rank(tool["Risk Level"]) > risk_rank(rec["highest"]):
            rec["highest"] = tool["Risk Level"]

    category_rows = []
    for cat, rec in sorted(category_map.items(), key=lambda i: i[1]["queries"], reverse=True):
        category_rows.append(
            {
                "Category": cat,
                "Total tools in category": len(rec["tools"]),
                "Total queries in category": rec["queries"],
                "% of all shadow traffic": round((rec["queries"] / total_shadow_queries) * 100, 2),
                "Highest risk tool in category": rec["highest"],
                "GDPR exposure count": rec["gdpr_count"],
            }
        )

    gdpr_tool_rows = [
        {
            "Tool": t["Tool Name"],
            "Category": t["Category"],
            "Risk": t["Risk Level"],
            "Queries": t["Total Queries (all sessions)"],
            "Unique IPs": t["Unique IPs"],
            "Risk Reasons": t["Risk Reasons (full list)"],
        }
        for t in master_tools
        if t["GDPR Concern"] == "YES"
    ]
    personal_data_exposure_est = sum((t["Queries"] * max(1, t["Unique IPs"])) for t in gdpr_tool_rows)

    alert_rows = []
    for idx, a in enumerate(alerts, start=1):
        alert_rows.append(
            {
                "#": idx,
                "Alert ID": a.get("id", a.get("alert_id", f"alert_{idx:03d}")),
                "Created At": a.get("created_at", "N/A"),
                "Tool Name": a.get("tool_name", "N/A"),
                "Session Date": a.get("session_date", "N/A"),
                "Severity": str(a.get("severity", "LOW")).upper(),
                "Status": str(a.get("status", "OPEN")).upper(),
                "Alert Message": a.get("message", a.get("alert_message", "N/A")),
                "Notes": a.get("notes", ""),
                "Resolved At": a.get("resolved_at", "N/A"),
            }
        )

    open_count = sum(1 for a in alert_rows if a["Status"] == "OPEN")
    ack_count = sum(1 for a in alert_rows if a["Status"] == "ACKNOWLEDGED")
    resolved_count = sum(1 for a in alert_rows if a["Status"] == "RESOLVED")
    high_count_alert = sum(1 for a in alert_rows if a["Severity"] == "HIGH")
    med_count_alert = sum(1 for a in alert_rows if a["Severity"] == "MEDIUM")
    low_count_alert = sum(1 for a in alert_rows if a["Severity"] == "LOW")

    durations = []
    for a in alert_rows:
        c = parse_dt(a["Created At"])
        r = parse_dt(a["Resolved At"])
        if c and r:
            durations.append((r - c).total_seconds() / 3600)
    avg_resolution = f"{mean(durations):.2f} hrs" if durations else "N/A"
    unresolved_high = [a for a in alert_rows if a["Status"] == "OPEN" and a["Severity"] == "HIGH"]

    catalog_categories = sorted({t.get("category", "N/A") for t in catalog_tools})
    catalog_mtime = "N/A"
    if CATALOG_PATH.exists():
        catalog_mtime = datetime.fromtimestamp(CATALOG_PATH.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S")
    settings_snapshot = [
        {"Setting": "Risk Threshold for Auto-Alerting", "Value": settings.get("risk_threshold", "N/A")},
        {"Setting": "Auto-Generate Alerts on Upload", "Value": settings.get("auto_alert", "N/A")},
        {"Setting": "Session Retention Limit", "Value": settings.get("session_limit", "N/A")},
        {"Setting": "SaaS Catalog — Total Tools", "Value": len(catalog_tools)},
        {"Setting": "SaaS Catalog — Categories", "Value": ", ".join(catalog_categories) or "N/A"},
        {"Setting": "SaaS Catalog — Last Modified", "Value": catalog_mtime},
        {"Setting": "App Version / Report Schema Version", "Value": "v1.0"},
    ]

    high_recs = []
    medium_recs = []
    gdpr_recs = []
    for t in master_tools:
        if t["Risk Level"] == "HIGH":
            ips = [d["IP Address"] for d in device_rows if t["Tool Name"] in d["GDPR Tools Accessed"] or d["Highest Risk Level"] == "HIGH"]
            high_recs.append(
                {
                    "priority": "HIGH",
                    "target": t["Tool Name"],
                    "text": (
                        f"BLOCK IMMEDIATELY | Category: {t['Category']} | Queries: {t['Total Queries (all sessions)']} "
                        f"across {t['Sessions Appeared In']} sessions from {t['Unique IPs']} devices | "
                        f"Risk Reasons: {t['Risk Reasons (full list)']} | GDPR: {t['GDPR Concern']} | "
                        f"Action: Block at DNS firewall level. Deploy {t['Recommended Alternative']} as replacement. "
                        f"Notify IT Security + device owners: {', '.join(ips[:10]) or 'N/A'}"
                    ),
                }
            )
        elif t["Risk Level"] == "MEDIUM":
            medium_recs.append(
                {
                    "priority": "MEDIUM",
                    "target": t["Tool Name"],
                    "text": (
                        f"REVIEW & POLICY DECISION REQUIRED | Category: {t['Category']} | Queries: {t['Total Queries (all sessions)']} "
                        f"| Devices: {t['Unique IPs']} | Risk Reasons: {t['Risk Reasons (full list)']} | "
                        f"Action: Schedule DLP review and determine business justification. "
                        f"Alternative: {t['Recommended Alternative']}"
                    ),
                }
            )
        if t["GDPR Concern"] == "YES":
            gdpr_recs.append(
                {
                    "priority": "HIGH",
                    "target": t["Tool Name"],
                    "text": (
                        f"GDPR DPIA REQUIRED | Personal data may be transmitted via detected domains. "
                        f"Accessed by {t['Unique IPs']} devices, {t['Total Queries (all sessions)']} queries. "
                        f"Action: Conduct DPIA immediately and notify DPO if processing continues."
                    ),
                }
            )

    device_actions = [
        {
            "priority": "MEDIUM",
            "target": row["IP Address"],
            "text": (
                f"USER AWARENESS REQUIRED | Device {row['IP Address']} | Highest Risk {row['Highest Risk Level']} | "
                f"Recommendation: IT to contact device owner and schedule security awareness training."
            ),
        }
        for row in repeat_offenders
    ]

    if overall_score >= 80:
        audit_freq = "weekly"
    elif overall_score >= 50:
        audit_freq = "bi-weekly"
    else:
        audit_freq = "monthly"
    systemic_recs = [
        "Harden DNS filtering policy with block-by-default for unapproved high-risk SaaS categories.",
        "Enable endpoint DLP controls on file upload and clipboard channels for unmanaged SaaS domains.",
        "Establish a SaaS governance workflow requiring security and procurement sign-off before adoption.",
        f"Run periodic DNS shadow IT audits {audit_freq} based on current exposure posture.",
    ]

    return {
        "title": report_title,
        "org_name": org_name,
        "generated_at": datetime.now(),
        "period_start": earliest,
        "period_end": latest,
        "sessions": normalized,
        "all_sessions": all_normalized,
        "executive": {
            "organisation": org_name,
            "generated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "period": f"{earliest} -> {latest}",
            "total_sessions": len(normalized),
            "total_queries": total_queries,
            "unique_domains": len(unique_domains),
            "unique_source_ips": len(unique_ips),
            "total_shadow_tools": len(tool_names),
            "overall_score": overall_score,
            "overall_verdict": overall_verdict,
            "overall_verdict_color": verdict_color,
            "highest": highest,
            "lowest": lowest,
            "trend": f"{trend_symbol} {trend_text}",
            "trend_symbol": trend_symbol,
            "trend_text": trend_text,
            "medium_threshold_exceeded": medium_threshold_exceeded,
            "new_high_tools": high_new_count,
            "total_cumulative_app_usage_time": format_duration(total_app_duration_sec),
            "average_session_duration": format_duration(avg_session_duration_sec),
            "longest_working_session": (
                f"{dt_text(longest_session['parsed_at'])} - {format_duration(int(longest_session.get('total_app_duration_sec', 0) or 0))}"
                if longest_session
                else "N/A"
            ),
            "shortest_working_session": (
                f"{dt_text(shortest_session['parsed_at'])} - {format_duration(int(shortest_session.get('total_app_duration_sec', 0) or 0))}"
                if shortest_session
                else "N/A"
            ),
            "most_active_hour_of_day": peak_hour,
            "total_pages_visited_all_sessions": int(sum(row["Page Visit Count"] for row in session_timing_rows)),
            "most_visited_page": most_visited_page,
            "total_shadow_app_usage_time": format_duration(total_shadow_app_usage_sec),
            "average_shadow_app_usage_per_session": format_duration(avg_shadow_usage_per_session_sec),
            "single_most_used_shadow_app": (
                f"{longest_total_tool['Tool Name']} - {longest_total_tool['Total Duration (all sessions)']}"
                if longest_total_tool
                else "N/A"
            ),
            "longest_uninterrupted_shadow_tool_session": (
                f"{longest_single_tool_session['duration_fmt']} - {longest_single_tool_session['tool_name']} - {dt_text(longest_single_tool_session['session_parsed_at'])}"
                if longest_single_tool_session
                else "N/A"
            ),
            "peak_shadow_activity_hour": shadow_peak_hour,
            "pct_workday_spent_on_shadow_tools": f"{pct_workday_shadow}%",
            "after_hours_shadow_usage_detected": (
                f"Yes - {len(after_hours_tool_names)} tools used after 18:00"
                if after_hours_tool_names
                else "No"
            ),
        },
        "devices": device_rows,
        "power_users": power_users,
        "gdpr_devices": gdpr_devices,
        "repeat_offenders": repeat_offenders,
        "master_tools": master_tools,
        "most_persistent": most_persistent,
        "fastest_growing": fastest_growing,
        "fastest_growth_delta": fastest_growth_delta if fastest_growth_delta > -10**9 else 0,
        "new_tools_latest": new_tools_latest,
        "disappeared_tools": disappeared_tools,
        "category_rows": category_rows,
        "gdpr_tool_rows": gdpr_tool_rows,
        "personal_data_exposure_est": personal_data_exposure_est,
        "alerts": alert_rows,
        "alerts_stats": {
            "Total Alerts Ever Created": len(alert_rows),
            "Currently Open": open_count,
            "Acknowledged": ack_count,
            "Resolved": resolved_count,
            "HIGH Severity": high_count_alert,
            "MEDIUM Severity": med_count_alert,
            "LOW Severity": low_count_alert,
            "Avg Time to Resolution": avg_resolution,
        },
        "unresolved_high": unresolved_high,
        "settings_snapshot": settings_snapshot,
        "duration_intelligence": {
            "master_table": duration_master_rows,
            "rankings": {
                "top_duration": top_duration_tools,
                "top_intensive": intensive_tools,
                "top_after_hours": after_hours_tools,
                "longest_total": longest_total_tool,
                "longest_single": longest_single_tool,
            },
            "timeline": duration_timeline,
        },
        "per_device_app_duration": per_device_app_rows,
        "device_usage_heatmap": ip_heatmap_rows,
        "tool_usage_heatmap": tool_heatmap_rows,
        "user_session_timing": {
            "rows": session_timing_rows,
            "summary": {
                "total_cumulative_time_sec": total_app_duration_sec,
                "total_cumulative_time_fmt": format_duration(total_app_duration_sec),
                "average_session_duration_sec": avg_session_duration_sec,
                "average_session_duration_fmt": format_duration(avg_session_duration_sec),
                "longest_session": longest_session,
                "shortest_session": shortest_session,
                "most_visited_page": most_visited_page,
                "most_visited_page_count": overall_page_counter.get(most_visited_page, 0) if most_visited_page != "N/A" else 0,
                "peak_usage_hour": peak_hour,
                "analysis_duration_total_sec": analysis_duration_sec_total,
                "analysis_duration_total_fmt": format_duration(analysis_duration_sec_total),
            },
        },
        "recommendations": {
            "high": high_recs,
            "medium": medium_recs,
            "gdpr": gdpr_recs,
            "device": device_actions,
            "systemic": systemic_recs,
        },
    }


def _xlsx_style_header(ws, row_idx):
    fill = PatternFill("solid", fgColor=GOLD)
    font = Font(color="000000", bold=True)
    for cell in ws[row_idx]:
        if cell.value is not None:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="left", vertical="center")


def _xlsx_auto_fit(ws):
    for idx, col in enumerate(ws.columns, start=1):
        max_len = 0
        col_letter = get_column_letter(idx)
        for cell in col:
            text = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(text))
            cell.font = cell.font.copy(color=CREAM)  # noqa: SLF001
        ws.column_dimensions[col_letter].width = max(15, min(50, max_len + 2))


def _sheet_background(ws):
    fill = PatternFill("solid", fgColor=PANEL_BG)
    for row in ws.iter_rows(min_row=1, max_row=max(200, ws.max_row + 30), min_col=1, max_col=max(15, ws.max_column)):
        for cell in row:
            if cell.value is None:
                cell.fill = fill
            else:
                if cell.fill.fill_type is None:
                    cell.fill = fill
                if cell.font is None:
                    cell.font = Font(color=CREAM)


def write_df(ws, df, start_row=1):
    if df.empty:
        ws.cell(row=start_row, column=1, value="No data")
        return start_row + 2
    headers = list(df.columns)
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=col_idx, value=h)
    for row_offset, row in enumerate(df.itertuples(index=False), start=1):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=start_row + row_offset, column=col_idx, value=value)
    _xlsx_style_header(ws, start_row)
    return start_row + len(df) + 2


def build_excel(report):
    wb = Workbook()
    wb.remove(wb.active)
    ts = report["generated_at"].strftime("%Y-%m-%d %H:%M:%S")

    def make_sheet(name):
        ws = wb.create_sheet(title=name)
        ws.sheet_view.showGridLines = False
        ws.oddHeader.left.text = report["org_name"]
        ws.oddHeader.right.text = ts
        return ws

    ws_cover = make_sheet("Cover")
    ws_cover.merge_cells("A1:F1")
    ws_cover["A1"] = report["title"]
    ws_cover["A1"].fill = PatternFill("solid", fgColor=GOLD)
    ws_cover["A1"].font = Font(color="000000", bold=True, size=18)
    ws_cover["A2"] = f"Organisation: {report['org_name']}"
    ws_cover["A3"] = f"Generated: {ts}"
    ws_cover["A4"] = f"Period: {report['period_start']} -> {report['period_end']}"
    ws_cover["A6"] = "Exposure Score"
    ws_cover["B6"] = report["executive"]["overall_score"]
    verdict = report["executive"]["overall_verdict"]
    ws_cover["C6"] = verdict
    risk_fill = PatternFill("solid", fgColor=RED if verdict in ("HIGH", "CRITICAL") else AMBER if verdict == "MODERATE" else GREEN)
    ws_cover["C6"].fill = risk_fill

    ws_exec = make_sheet("Executive Summary")
    exec_rows = [
        ("Organisation name", report["executive"]["organisation"]),
        ("Report generated", report["executive"]["generated"]),
        ("Report period", report["executive"]["period"]),
        ("Total sessions analysed", report["executive"]["total_sessions"]),
        ("Total DNS queries analysed", report["executive"]["total_queries"]),
        ("Total unique domains seen", report["executive"]["unique_domains"]),
        ("Total unique source IPs", report["executive"]["unique_source_ips"]),
        ("Total shadow tools ever detected", report["executive"]["total_shadow_tools"]),
        ("Overall system Exposure Score", report["executive"]["overall_score"]),
        ("Overall risk verdict", report["executive"]["overall_verdict"]),
        ("Highest session", f"{report['executive']['highest']['parsed_at'] if report['executive']['highest'] else 'N/A'} / {report['executive']['highest']['score'] if report['executive']['highest'] else 'N/A'}"),
        ("Lowest session", f"{report['executive']['lowest']['parsed_at'] if report['executive']['lowest'] else 'N/A'} / {report['executive']['lowest']['score'] if report['executive']['lowest'] else 'N/A'}"),
        ("Score trend", report["executive"]["trend"]),
        ("Sessions above medium threshold", report["executive"]["medium_threshold_exceeded"]),
        ("New high risk tools first detected", report["executive"]["new_high_tools"]),
        ("Total cumulative app usage time", report["executive"]["total_cumulative_app_usage_time"]),
        ("Average session duration", report["executive"]["average_session_duration"]),
        ("Longest working session", report["executive"]["longest_working_session"]),
        ("Most active hour of day", report["executive"]["most_active_hour_of_day"]),
        ("Total pages visited (all sessions)", report["executive"]["total_pages_visited_all_sessions"]),
        ("Most visited page", report["executive"]["most_visited_page"]),
        ("Total shadow app usage time (all tools, all sessions)", report["executive"]["total_shadow_app_usage_time"]),
        ("Average shadow app usage per session", report["executive"]["average_shadow_app_usage_per_session"]),
        ("Single most used shadow app", report["executive"]["single_most_used_shadow_app"]),
        ("Longest uninterrupted shadow tool session", report["executive"]["longest_uninterrupted_shadow_tool_session"]),
        ("Peak shadow activity hour", report["executive"]["peak_shadow_activity_hour"]),
        ("% of workday spent on shadow tools (est.)", report["executive"]["pct_workday_spent_on_shadow_tools"]),
        ("After-hours shadow usage detected", report["executive"]["after_hours_shadow_usage_detected"]),
    ]
    for i, (k, v) in enumerate(exec_rows, start=1):
        ws_exec.cell(row=i, column=1, value=k)
        ws_exec.cell(row=i, column=2, value=v)
    for i in range(1, len(exec_rows) + 1):
        ws_exec.cell(row=i, column=1).fill = PatternFill("solid", fgColor=GOLD)
        ws_exec.cell(row=i, column=1).font = Font(color="000000", bold=True)

    ws_sessions = make_sheet("Sessions Overview")
    sess_df = pd.DataFrame(
        [
            {
                "Session ID": s["session_id"],
                "Parsed At": s["parsed_at"],
                "Source File": s["source_file"],
                "File Size": s["file_size"],
                "Total DNS Queries": s["total_queries"],
                "Unique Domains": s["unique_domains"],
                "Unique Source IPs": s["unique_source_ips"],
                "Shadow Tools": s["shadow_tools"],
                "HIGH": s["high_count"],
                "MEDIUM": s["medium_count"],
                "LOW": s["low_count"],
                "GDPR": s["gdpr_count"],
                "Exposure Score": s["score"],
                "Verdict": s["score_label"],
                "Auto-alerts": s["auto_alerts"],
            }
            for s in report["sessions"]
        ]
    )
    write_df(ws_sessions, sess_df)
    ws_sessions.freeze_panes = "A2"
    ws_sessions.auto_filter.ref = ws_sessions.dimensions

    for idx, s in enumerate(report["sessions"], start=1):
        ws_s = make_sheet(f"Session_{idx}")
        meta_df = pd.DataFrame(
            [
                ("Session ID", s["session_id"]),
                ("Parsed At", s["parsed_at"]),
                ("Source File", s["source_file"]),
                ("File Size", s["file_size"]),
                ("Total DNS Queries", s["total_queries"]),
                ("Unique Domains", s["unique_domains"]),
                ("Unique Source IPs", s["unique_source_ips"]),
                ("Shadow Tools", s["shadow_tools"]),
                ("Exposure Score", s["score"]),
                ("App Opened At", dt_text(s.get("app_opened_at"))),
                ("Analysis Started At", dt_text(s.get("analysis_started_at"))),
                ("Analysis Ended At", dt_text(s.get("analysis_ended_at"))),
                ("App Closed At", dt_text(s.get("app_closed_at"))),
                ("Analysis Duration", s.get("analysis_duration_fmt", "N/A")),
                ("Total App Session Duration", s.get("total_app_duration_fmt", "N/A")),
                ("Pages Visited", pages_text(s.get("pages_visited", []))),
                ("Page Visit Count", s.get("page_visit_count", 0)),
            ],
            columns=["Field", "Value"],
        )
        row = write_df(ws_s, meta_df, start_row=1)
        tools_df = pd.DataFrame(
            [
                {
                    "Tool Name": t["tool_name"],
                    "Category": t["category"],
                    "Risk": t["risk"],
                    "Query Count": t["query_count"],
                    "First Seen": dt_text((t.get("duration") or {}).get("first_query_at")),
                    "Last Seen": dt_text((t.get("duration") or {}).get("last_query_at")),
                    "Total Span": (t.get("duration") or {}).get("total_span_fmt", "00h 00m 00s"),
                    "Active Minutes": int((t.get("duration") or {}).get("active_minutes", 0) or 0),
                    "Queries/Min": float((t.get("duration") or {}).get("queries_per_minute", 0) or 0),
                    "Peak Hour": (
                        f"{int((t.get('duration') or {}).get('peak_usage_hour', 0)):02d}:00-{(int((t.get('duration') or {}).get('peak_usage_hour', 0)) + 1) % 24:02d}:00"
                        if (t.get("duration") or {}).get("peak_usage_hour") is not None
                        else "N/A"
                    ),
                    "% of Total Queries": round((t["query_count"] / max(1, s["total_queries"])) * 100, 2),
                    "Unique IPs": t["unique_ips"],
                    "GDPR": "YES" if t["gdpr"] else "No",
                    "Risk Reasons": "; ".join(t["risk_reasons"]) or "N/A",
                    "Recommended Alternative": t["alternative"],
                }
                for t in s["tools"]
            ]
        )
        ws_s.cell(row=row, column=1, value="Shadow Tools")
        row = write_df(ws_s, tools_df, start_row=row + 1)
        ip_df = pd.DataFrame(
            [
                {
                    "IP Address": ip["ip"],
                    "Query Count": ip["query_count"],
                    "Tools Accessed": ", ".join(ip["tools_accessed"]) or "N/A",
                    "Risk Level of Tools": ip["risk_level"],
                }
                for ip in s["ip_rows"]
            ]
        )
        ws_s.cell(row=row, column=1, value="Source IPs")
        write_df(ws_s, ip_df, start_row=row + 1)
        ws_s.freeze_panes = "A2"
        ws_s.auto_filter.ref = ws_s.dimensions

    ws_devices = make_sheet("All Devices")
    dev_df = pd.DataFrame(report["devices"])
    if not dev_df.empty and "repeat_offender" in dev_df.columns:
        dev_df = dev_df.drop(columns=["repeat_offender"])
    write_df(ws_devices, dev_df)
    ws_devices.freeze_panes = "A2"
    ws_devices.auto_filter.ref = ws_devices.dimensions

    ws_tools = make_sheet("Master Tool Table")
    m_df = pd.DataFrame([{k: v for k, v in t.items() if not k.startswith("_")} for t in report["master_tools"]])
    write_df(ws_tools, m_df)
    ws_tools.freeze_panes = "A2"
    ws_tools.auto_filter.ref = ws_tools.dimensions

    ws_gdpr = make_sheet("GDPR Exposure")
    gdpr_df = pd.DataFrame(report["gdpr_tool_rows"])
    write_df(ws_gdpr, gdpr_df)
    ws_gdpr.freeze_panes = "A2"
    ws_gdpr.auto_filter.ref = ws_gdpr.dimensions

    ws_alert = make_sheet("Alerts Log")
    alert_df = pd.DataFrame(report["alerts"])
    write_df(ws_alert, alert_df)
    ws_alert.freeze_panes = "A2"
    ws_alert.auto_filter.ref = ws_alert.dimensions

    ws_rec = make_sheet("Recommendations")
    rec_rows = report["recommendations"]["high"] + report["recommendations"]["medium"] + report["recommendations"]["gdpr"] + report["recommendations"]["device"]
    rec_df = pd.DataFrame(
        [{"Priority": r["priority"], "Tool/Device": r["target"], "Full Action Text": r["text"]} for r in rec_rows]
    )
    write_df(ws_rec, rec_df)
    ws_rec.freeze_panes = "A2"
    ws_rec.auto_filter.ref = ws_rec.dimensions

    ws_timing = make_sheet("User Session Timing")
    timing_df = pd.DataFrame(report["user_session_timing"]["rows"])
    timing_view_df = timing_df.copy()
    if not timing_view_df.empty:
        for col_name in ["Analysis Duration Sec", "Total App Duration Sec"]:
            if col_name in timing_view_df.columns:
                timing_view_df = timing_view_df.drop(columns=[col_name])
    last_row = write_df(ws_timing, timing_view_df, start_row=1)

    summary = report["user_session_timing"]["summary"]
    summary_rows = [
        ("Total cumulative time across all sessions", summary["total_cumulative_time_fmt"]),
        ("Average session duration", summary["average_session_duration_fmt"]),
        (
            "Longest session",
            (
                f"{dt_text(summary['longest_session']['parsed_at'])} - "
                f"{format_duration(int(summary['longest_session'].get('total_app_duration_sec', 0) or 0))}"
                if summary.get("longest_session")
                else "N/A"
            ),
        ),
    ]
    ws_timing.cell(row=last_row, column=1, value="Metric")
    ws_timing.cell(row=last_row, column=2, value="Value")
    _xlsx_style_header(ws_timing, last_row)
    for idx, (metric, value) in enumerate(summary_rows, start=1):
        ws_timing.cell(row=last_row + idx, column=1, value=metric)
        ws_timing.cell(row=last_row + idx, column=2, value=value)
        ws_timing.cell(row=last_row + idx, column=1).fill = PatternFill("solid", fgColor=GOLD)
        ws_timing.cell(row=last_row + idx, column=1).font = Font(color="000000", bold=True)
        if metric == "Longest session":
            ws_timing.cell(row=last_row + idx, column=2).fill = PatternFill("solid", fgColor=AMBER)
            ws_timing.cell(row=last_row + idx, column=2).font = Font(color="000000", bold=True)
    ws_timing.freeze_panes = "A2"
    ws_timing.auto_filter.ref = ws_timing.dimensions

    ws_dur_master = make_sheet("App Duration Master")
    dmaster_df = pd.DataFrame(
        [
            {
                "Tool": row.get("Tool Name"),
                "Category": row.get("Category"),
                "Risk": row.get("Risk Level"),
                "First Ever Seen": row.get("First Detected"),
                "Last Ever Seen": row.get("Last Detected"),
                "Total Duration": row.get("Total Duration (all sessions)"),
                "Total Duration Sec": row.get("Total Duration Sec", 0),
                "Avg Per Session": row.get("Avg Duration Per Session"),
                "Longest Session": row.get("Longest Single Session Duration"),
                "Active Minutes Total": row.get("Total Active Minutes", 0),
                "Total Queries": row.get("Total Queries (all sessions)", 0),
                "Peak Hour": row.get("Peak Usage Hour", "N/A"),
            }
            for row in report["duration_intelligence"]["master_table"]
        ]
    )
    dmaster_row = write_df(ws_dur_master, dmaster_df.drop(columns=["Total Duration Sec"], errors="ignore"), start_row=1)
    if not dmaster_df.empty:
        max_sec = int(dmaster_df["Total Duration Sec"].max() or 0)
        min_sec = int(dmaster_df["Total Duration Sec"].min() or 0)
        for idx, rec in enumerate(dmaster_df.to_dict("records"), start=2):
            sec = int(rec.get("Total Duration Sec", 0) or 0)
            risk_ratio = 0 if max_sec == min_sec else (sec - min_sec) / (max_sec - min_sec)
            if risk_ratio > 0.66:
                color = "F44336"
            elif risk_ratio > 0.33:
                color = "FF9800"
            else:
                color = "4CAF50"
            ws_dur_master.cell(row=idx, column=6).fill = PatternFill("solid", fgColor=color)
        longest_idx = dmaster_df["Total Duration Sec"].idxmax() + 2
        ws_dur_master.cell(row=longest_idx, column=1).border = Border(left=Side(style="thick", color=GOLD))

    total_dur = report["duration_intelligence"]["master_table"]
    total_dur_sec = int(sum(int(x.get("Total Duration Sec", 0) or 0) for x in total_dur))
    ws_dur_master.cell(row=dmaster_row, column=1, value="Total Duration Across All Tools")
    ws_dur_master.cell(row=dmaster_row, column=2, value=format_duration(total_dur_sec))
    ws_dur_master.cell(row=dmaster_row, column=1).fill = PatternFill("solid", fgColor=GOLD)
    ws_dur_master.cell(row=dmaster_row, column=2).fill = PatternFill("solid", fgColor=GOLD)
    ws_dur_master.cell(row=dmaster_row, column=1).font = Font(color="000000", bold=True)
    ws_dur_master.cell(row=dmaster_row, column=2).font = Font(color="000000", bold=True)
    ws_dur_master.freeze_panes = "A2"
    ws_dur_master.auto_filter.ref = ws_dur_master.dimensions

    ws_ip_app = make_sheet("Per-IP Per-App Duration")
    ip_app_df = pd.DataFrame(report["per_device_app_duration"])
    ip_app_view = ip_app_df.drop(columns=["Total Duration Sec"], errors="ignore") if not ip_app_df.empty else ip_app_df
    write_df(ws_ip_app, ip_app_view, start_row=1)
    if not ip_app_view.empty:
        risk_col = list(ip_app_view.columns).index("Risk") + 1 if "Risk" in ip_app_view.columns else None
        duration_col = list(ip_app_view.columns).index("Total Duration") + 1 if "Total Duration" in ip_app_view.columns else None
        for row_idx in range(2, len(ip_app_view) + 2):
            if risk_col:
                risk_val = str(ws_ip_app.cell(row=row_idx, column=risk_col).value or "").upper()
                fill = "4CAF50"
                if risk_val == "HIGH":
                    fill = "F44336"
                elif risk_val == "MEDIUM":
                    fill = "FF9800"
                ws_ip_app.cell(row=row_idx, column=risk_col).fill = PatternFill("solid", fgColor=fill)
            if duration_col and not ip_app_df.empty:
                dsec = int(ip_app_df.iloc[row_idx - 2].get("Total Duration Sec", 0) or 0)
                if dsec > 8 * 3600:
                    ws_ip_app.cell(row=row_idx, column=duration_col).fill = PatternFill("solid", fgColor="F44336")
                elif dsec > 4 * 3600:
                    ws_ip_app.cell(row=row_idx, column=duration_col).fill = PatternFill("solid", fgColor="FF9800")
    ws_ip_app.freeze_panes = "A2"
    ws_ip_app.auto_filter.ref = ws_ip_app.dimensions

    ws_heat = make_sheet("Hourly Usage Heatmap")
    heat_df = pd.DataFrame(report["tool_usage_heatmap"])
    write_df(ws_heat, heat_df, start_row=1)
    if not heat_df.empty:
        max_heat = int(heat_df[[c for c in heat_df.columns if c.startswith("Hour ")]].max().max() or 0)
        risk_map = {row.get("Tool Name"): row.get("Risk Level") for row in report["master_tools"]}
        for ridx, rec in enumerate(heat_df.to_dict("records"), start=2):
            tool_name = rec.get("Tool")
            risk = str(risk_map.get(tool_name, "LOW")).upper()
            border_color = GREEN
            if risk == "HIGH":
                border_color = RED
            elif risk == "MEDIUM":
                border_color = AMBER
            ws_heat.cell(row=ridx, column=1).border = Border(left=Side(style="thick", color=border_color))
            for cidx, col in enumerate(heat_df.columns, start=1):
                if not str(col).startswith("Hour "):
                    continue
                val = int(rec.get(col, 0) or 0)
                ratio = 0 if max_heat == 0 else (val / max_heat)
                shade = int(255 - (ratio * 180))
                hex_color = f"FF{shade:02X}{shade:02X}"
                ws_heat.cell(row=ridx, column=cidx).fill = PatternFill("solid", fgColor=hex_color)
    ws_heat.freeze_panes = "A2"
    ws_heat.auto_filter.ref = ws_heat.dimensions

    ws_set = make_sheet("Settings Snapshot")
    set_df = pd.DataFrame(report["settings_snapshot"])
    write_df(ws_set, set_df)
    ws_set.freeze_panes = "A2"
    ws_set.auto_filter.ref = ws_set.dimensions

    for ws in wb.worksheets:
        _sheet_background(ws)
        _xlsx_auto_fit(ws)

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


def build_pdf(report):
    buff = BytesIO()
    doc = SimpleDocTemplate(buff, pagesize=A4, leftMargin=16 * mm, rightMargin=16 * mm, topMargin=16 * mm, bottomMargin=16 * mm)
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Heading1"], textColor=colors.HexColor("#FFD700"), fontSize=24, leading=30)
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], textColor=colors.HexColor("#FFD700"), fontSize=16, leading=20)
    body = ParagraphStyle("body", parent=styles["BodyText"], textColor=colors.HexColor("#E8D48A"), fontSize=10, leading=14)

    elements = []
    elements.append(Spacer(1, 10 * mm))
    elements.append(Paragraph("SHADOW IT DNA MAP", h1))
    elements.append(Paragraph("Full System Intelligence Report", h2))
    elements.append(Spacer(1, 4 * mm))
    elements.append(Paragraph(f"Organisation: {report['org_name']}", body))
    elements.append(Paragraph(f"Report Period: {report['period_start']} -> {report['period_end']}", body))
    elements.append(Paragraph(f"Generated: {report['generated_at'].strftime('%Y-%m-%d %H:%M:%S')}", body))

    drawing = Drawing(120, 120)
    verdict = report["executive"]["overall_verdict"]
    ring_color = colors.HexColor("#F44336") if verdict in ("HIGH", "CRITICAL") else colors.HexColor("#FF9800") if verdict == "MODERATE" else colors.HexColor("#4CAF50")
    drawing.add(Circle(60, 60, 45, strokeWidth=6, strokeColor=ring_color, fillColor=None))
    drawing.add(String(48, 56, str(int(report["executive"]["overall_score"])), fontSize=16, fillColor=ring_color))
    elements.append(Spacer(1, 4 * mm))
    elements.append(drawing)
    elements.append(Paragraph(f"Risk Verdict: {verdict}", h2))
    elements.append(Spacer(1, 8 * mm))

    elements.append(Paragraph("Table of Contents", h2))
    toc_rows = [
        ["A. Executive Summary", ""],
        ["A2. User Working Session Timeline", ""],
        ["B. Per-Session Breakdown", ""],
        ["C. All Users / Devices", ""],
        ["D. Shadow Tool Intelligence", ""],
        ["E. Alerts History", ""],
        ["F. Settings Snapshot", ""],
        ["G. Recommendations", ""],
    ]
    toc_tbl = Table(toc_rows, colWidths=[130 * mm, 35 * mm])
    toc_tbl.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#13100A")),
                ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#E8D48A")),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#2A2200")),
            ]
        )
    )
    elements.append(toc_tbl)
    elements.append(Spacer(1, 8 * mm))

    elements.append(Paragraph("A. Executive Summary", h2))
    exec_data = [["Field", "Value"]] + [[k, str(v)] for k, v in report["executive"].items() if k not in {"highest", "lowest", "overall_verdict_color", "trend_symbol", "trend_text"}]
    exec_table = Table(exec_data, colWidths=[70 * mm, 95 * mm])
    exec_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FFD700")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#13100A")),
                ("TEXTCOLOR", (0, 1), (-1, -1), colors.HexColor("#E8D48A")),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#2A2200")),
            ]
        )
    )
    elements.append(exec_table)
    elements.append(Spacer(1, 6 * mm))

    elements.append(PageBreak())
    elements.append(Paragraph("⏱ User Working Session Timeline", h2))
    timing_rows = report["user_session_timing"]["rows"]
    table1 = [["Session", "App Start", "Analysis Start", "Analysis End", "App End", "Analysis Duration", "Total Duration"]]
    for row in timing_rows:
        table1.append(
            [
                str(row["Session #"]),
                row["App Opened At"],
                row["Analysis Start"],
                row["Analysis End"],
                row["App Closed At"],
                row["Analysis Duration"],
                row["Total App Duration"],
            ]
        )
    timing_table = Table(table1[:80], colWidths=[12 * mm, 24 * mm, 24 * mm, 24 * mm, 24 * mm, 26 * mm, 26 * mm])
    timing_style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FFD700")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
        ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#13100A")),
        ("TEXTCOLOR", (0, 1), (-1, -1), colors.HexColor("#E8D48A")),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#2A2200")),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
    ]
    for idx, row in enumerate(timing_rows, start=1):
        dur_sec = int(row.get("Total App Duration Sec", 0) or 0)
        color = colors.HexColor("#4CAF50")
        if 900 <= dur_sec <= 3600:
            color = colors.HexColor("#FF9800")
        elif dur_sec > 3600:
            color = colors.HexColor("#F44336")
        timing_style.append(("TEXTCOLOR", (6, idx), (6, idx), color))
        timing_style.append(("FONTNAME", (6, idx), (6, idx), "Helvetica-Bold"))
    timing_table.setStyle(TableStyle(timing_style))
    elements.append(Paragraph("Table 1 — Session Timing Log", body))
    elements.append(timing_table)
    elements.append(Spacer(1, 4 * mm))

    timing_summary = report["user_session_timing"]["summary"]
    table2 = [
        ["Metric", "Value"],
        ["Total cumulative time", timing_summary["total_cumulative_time_fmt"]],
        ["Average session duration", timing_summary["average_session_duration_fmt"]],
        [
            "Longest session",
            (
                f"{dt_text(timing_summary['longest_session']['parsed_at'])} - "
                f"{format_duration(int(timing_summary['longest_session'].get('total_app_duration_sec', 0) or 0))}"
                if timing_summary.get("longest_session")
                else "N/A"
            ),
        ],
        [
            "Shortest session",
            (
                f"{dt_text(timing_summary['shortest_session']['parsed_at'])} - "
                f"{format_duration(int(timing_summary['shortest_session'].get('total_app_duration_sec', 0) or 0))}"
                if timing_summary.get("shortest_session")
                else "N/A"
            ),
        ],
        ["Most visited page", f"{timing_summary['most_visited_page']} ({timing_summary['most_visited_page_count']})"],
    ]
    summary_table = Table(table2, colWidths=[70 * mm, 95 * mm])
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FFD700")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#13100A")),
                ("TEXTCOLOR", (0, 1), (-1, -1), colors.HexColor("#E8D48A")),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#2A2200")),
            ]
        )
    )
    elements.append(Paragraph("Table 2 — Cumulative Usage Summary", body))
    elements.append(summary_table)
    elements.append(Spacer(1, 4 * mm))

    # Timeline visual: grey = total app session, gold = active analysis window.
    timeline = Drawing(520, 170)
    timeline.add(String(0, 155, "Timeline Visual", fontSize=10, fillColor=colors.HexColor("#FFD700")))
    min_hour = 0
    max_hour = 24
    valid_rows = [r for r in timing_rows if parse_dt(r["App Opened At"]) and parse_dt(r["Analysis End"])]
    if valid_rows:
        starts = [parse_dt(r["App Opened At"]) for r in valid_rows]
        ends = [parse_dt(r["App Closed At"]) or parse_dt(r["Analysis End"]) for r in valid_rows]
        min_hour = min(dt.hour for dt in starts)
        max_hour = max(dt.hour for dt in ends) + 1
        if max_hour <= min_hour:
            max_hour = min_hour + 1

    span = max(1, max_hour - min_hour)
    x0 = 75
    width = 420
    y_start = 132
    step = 18
    timeline.add(Line(x0, 20, x0 + width, 20, strokeColor=colors.HexColor("#E8D48A"), strokeWidth=1))
    for h in range(min_hour, max_hour + 1):
        x = x0 + ((h - min_hour) / span) * width
        timeline.add(Line(x, 17, x, 23, strokeColor=colors.HexColor("#E8D48A"), strokeWidth=1))
        timeline.add(String(x - 8, 5, f"{h:02d}", fontSize=7, fillColor=colors.HexColor("#E8D48A")))

    for idx, row in enumerate(valid_rows[:7], start=1):
        app_open = parse_dt(row["App Opened At"])
        analysis_start = parse_dt(row["Analysis Start"]) or app_open
        analysis_end = parse_dt(row["Analysis End"]) or analysis_start
        app_end = parse_dt(row["App Closed At"]) or analysis_end
        y = y_start - (idx - 1) * step
        if y < 28:
            break

        def x_for(dt_value):
            total = dt_value.hour + (dt_value.minute / 60.0) + (dt_value.second / 3600.0)
            return x0 + ((total - min_hour) / span) * width

        x_app_open = x_for(app_open)
        x_analysis_start = x_for(analysis_start)
        x_analysis_end = x_for(analysis_end)
        x_app_end = x_for(app_end)

        timeline.add(String(5, y - 2, f"S{row['Session #']}", fontSize=8, fillColor=colors.HexColor("#E8D48A")))
        timeline.add(
            Rect(
                x_app_open,
                y - 5,
                max(1, x_app_end - x_app_open),
                8,
                strokeWidth=0,
                fillColor=colors.HexColor("#6E6E6E"),
            )
        )
        timeline.add(
            Rect(
                x_analysis_start,
                y - 5,
                max(1, x_analysis_end - x_analysis_start),
                8,
                strokeWidth=0,
                fillColor=colors.HexColor("#FFD700"),
            )
        )
    elements.append(timeline)
    elements.append(Spacer(1, 6 * mm))

    elements.append(PageBreak())
    elements.append(Paragraph("Shadow App Duration Intelligence", h2))
    d_master = report["duration_intelligence"]["master_table"]
    d_rows = [["Tool", "Category", "Risk", "Total Duration", "Avg/Session", "Longest", "QPM", "After Hours"]]
    for row in d_master[:40]:
        d_rows.append(
            [
                row.get("Tool Name", "N/A"),
                row.get("Category", "N/A"),
                row.get("Risk Level", "N/A"),
                row.get("Total Duration (all sessions)", "00h 00m 00s"),
                row.get("Avg Duration Per Session", "00h 00m 00s"),
                row.get("Longest Single Session Duration", "00h 00m 00s"),
                str(row.get("Avg Queries/Min", 0)),
                f"{row.get('After Hours %', 0)}% {row.get('After Hours Flag', '')}",
            ]
        )
    d_table = Table(d_rows, colWidths=[34 * mm, 24 * mm, 14 * mm, 24 * mm, 22 * mm, 22 * mm, 12 * mm, 28 * mm])
    d_style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FFD700")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
        ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#13100A")),
        ("TEXTCOLOR", (0, 1), (-1, -1), colors.HexColor("#E8D48A")),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#2A2200")),
        ("FONTNAME", (3, 1), (3, -1), "Helvetica-Bold"),
    ]
    d_table.setStyle(TableStyle(d_style))
    elements.append(Paragraph("Table 1 — App Duration Master Table", body))
    elements.append(d_table)
    elements.append(Spacer(1, 3 * mm))

    ranks = report["duration_intelligence"]["rankings"]
    rank_rows = [["Ranking", "Result"]]
    for idx, row in enumerate(ranks.get("top_duration", [])[:5], start=1):
        rank_rows.append([f"Top Duration #{idx}", f"{row['Tool Name']} - {row['Total Duration (all sessions)']}"])
    for idx, row in enumerate(ranks.get("top_intensive", [])[:5], start=1):
        rank_rows.append([f"Most Intensive #{idx}", f"{row['Tool Name']} - {row['Avg Queries/Min']} qpm"])
    for idx, row in enumerate(ranks.get("top_after_hours", [])[:5], start=1):
        rank_rows.append([f"After-Hours #{idx}", f"{row['Tool Name']} - {row['After Hours %']}%"])
    rank_table = Table(rank_rows, colWidths=[52 * mm, 113 * mm])
    rank_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FFD700")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#13100A")),
                ("TEXTCOLOR", (0, 1), (-1, -1), colors.HexColor("#E8D48A")),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#2A2200")),
            ]
        )
    )
    elements.append(Paragraph("Table 2 — Duration Rankings", body))
    elements.append(rank_table)
    elements.append(Spacer(1, 3 * mm))

    # Visual 1: Horizontal duration bars by risk color.
    bars = Drawing(520, 160)
    bars.add(String(0, 145, "Duration Bar Chart", fontSize=10, fillColor=colors.HexColor("#FFD700")))
    top_tools = sorted(d_master, key=lambda x: x.get("Total Duration Sec", 0), reverse=True)[:7]
    max_dur = max([int(t.get("Total Duration Sec", 0) or 0) for t in top_tools], default=1)
    for i, row in enumerate(top_tools, start=1):
        y = 140 - i * 18
        risk = str(row.get("Risk Level", "LOW")).upper()
        color = colors.HexColor("#4CAF50")
        if risk == "HIGH":
            color = colors.HexColor("#F44336")
        elif risk == "MEDIUM":
            color = colors.HexColor("#FF9800")
        width = int((int(row.get("Total Duration Sec", 0) or 0) / max_dur) * 300)
        bars.add(String(0, y, row.get("Tool Name", "N/A"), fontSize=8, fillColor=colors.HexColor("#E8D48A")))
        bars.add(Rect(110, y - 2, max(1, width), 8, strokeWidth=0, fillColor=color))
        bars.add(String(420, y, row.get("Total Duration (all sessions)", "00h 00m 00s"), fontSize=8, fillColor=colors.HexColor("#E8D48A")))
    elements.append(bars)
    elements.append(Spacer(1, 2 * mm))

    # Visual 2: 24-hour activity clock.
    clock = Drawing(220, 220)
    cx, cy, r = 110, 110, 70
    hour_map = Counter()
    for row in report.get("tool_usage_heatmap", []):
        for h in range(24):
            hour_map[h] += int(row.get(f"Hour {h:02d}", 0) or 0)
    max_hour = max(hour_map.values(), default=1)
    clock.add(Circle(cx, cy, r, strokeColor=colors.HexColor("#E8D48A"), strokeWidth=1, fillColor=None))
    for h in range(24):
        intensity = int((hour_map.get(h, 0) / max_hour) * 255) if max_hour else 0
        shade = max(0, min(255, 255 - intensity))
        fill = colors.Color(1, shade / 255, shade / 255)
        angle = (h / 24) * 360
        x = cx + (r - 8) * math.cos(math.radians(angle - 90))
        y = cy + (r - 8) * math.sin(math.radians(angle - 90))
        clock.add(Circle(x, y, 6, strokeWidth=0, fillColor=fill))
    clock.add(String(cx - 80, 200, "Shadow Activity by Hour of Day", fontSize=9, fillColor=colors.HexColor("#FFD700")))
    clock.add(String(cx - 4, cy + r + 8, "0", fontSize=8, fillColor=colors.HexColor("#FFD700")))
    clock.add(String(cx + r + 6, cy - 2, "6", fontSize=8, fillColor=colors.HexColor("#FFD700")))
    clock.add(String(cx - 8, cy - r - 14, "12", fontSize=8, fillColor=colors.HexColor("#FFD700")))
    clock.add(String(cx - r - 14, cy - 2, "18", fontSize=8, fillColor=colors.HexColor("#FFD700")))
    elements.append(clock)

    elements.append(PageBreak())
    elements.append(Paragraph("Per-Device App Duration", h2))
    ip_rows = [["IP", "Tool", "Category", "Risk", "First", "Last", "Duration", "Active Min", "Queries", "After Hours"]]
    for row in report["per_device_app_duration"][:70]:
        first_dt = parse_dt(row.get("First Seen"))
        last_dt = parse_dt(row.get("Last Seen"))
        after_hours_flag = "🌙" if (first_dt and (first_dt.hour < 9 or first_dt.hour >= 18)) or (last_dt and (last_dt.hour < 9 or last_dt.hour >= 18)) else ""
        ip_rows.append(
            [
                row.get("IP Address", "N/A"),
                row.get("Tool Name", "N/A"),
                row.get("Category", "N/A"),
                row.get("Risk", "N/A"),
                row.get("First Seen", "N/A"),
                row.get("Last Seen", "N/A"),
                row.get("Total Duration", "00h 00m 00s"),
                str(row.get("Active Minutes", 0)),
                str(row.get("Query Count", 0)),
                after_hours_flag,
            ]
        )
    ip_table = Table(ip_rows, colWidths=[20 * mm, 22 * mm, 20 * mm, 13 * mm, 22 * mm, 22 * mm, 20 * mm, 12 * mm, 12 * mm, 10 * mm])
    ip_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FFD700")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#13100A")),
                ("TEXTCOLOR", (0, 1), (-1, -1), colors.HexColor("#E8D48A")),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#2A2200")),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
            ]
        )
    )
    elements.append(ip_table)
    elements.append(Spacer(1, 6 * mm))

    elements.append(Paragraph("B. Per-Session Breakdown", h2))
    for idx, s in enumerate(report["sessions"], start=1):
        elements.append(Paragraph(f"Session {idx}: {s['parsed_at']}", body))
        session_data = [
            ["Field", "Value"],
            ["Session ID", s["session_id"]],
            ["Source File", s["source_file"]],
            ["Total DNS Queries", s["total_queries"]],
            ["Unique Domains", s["unique_domains"]],
            ["Unique Source IPs", s["unique_source_ips"]],
            ["Shadow Tools", s["shadow_tools"]],
            ["Exposure Score", s["score"]],
        ]
        stbl = Table(session_data, colWidths=[70 * mm, 95 * mm])
        stbl.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FFD700")),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#13100A")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                    ("TEXTCOLOR", (0, 1), (-1, -1), colors.HexColor("#E8D48A")),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#2A2200")),
                ]
            )
        )
        elements.append(stbl)
        elements.append(Spacer(1, 3 * mm))

    elements.append(Paragraph("C. Device Risk Roster", h2))
    dev_rows = [["IP", "Queries", "Sessions", "Tools", "Highest Risk", "GDPR Tools"]]
    for r in report["devices"]:
        dev_rows.append(
            [
                r["IP Address"],
                str(r["Total Queries (all sessions)"]),
                str(r["Sessions Active"]),
                str(r["Shadow Tools Used"]),
                r["Highest Risk Level"],
                r["GDPR Tools Accessed"],
            ]
        )
    dev_table = Table(dev_rows[:60], colWidths=[25 * mm, 22 * mm, 18 * mm, 18 * mm, 20 * mm, 62 * mm])
    dev_table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FFD700")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.black), ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#13100A")), ("TEXTCOLOR", (0, 1), (-1, -1), colors.HexColor("#E8D48A")), ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#2A2200"))]))
    elements.append(dev_table)
    elements.append(Spacer(1, 6 * mm))

    elements.append(Paragraph("D. Master Tool Intelligence", h2))
    tool_rows = [["Tool", "Category", "Risk", "Queries", "Sessions", "GDPR"]]
    for t in report["master_tools"]:
        tool_rows.append([t["Tool Name"], t["Category"], t["Risk Level"], str(t["Total Queries (all sessions)"]), str(t["Sessions Appeared In"]), t["GDPR Concern"]])
    tool_table = Table(tool_rows[:80], colWidths=[55 * mm, 35 * mm, 20 * mm, 22 * mm, 18 * mm, 18 * mm])
    tool_table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FFD700")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.black), ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#13100A")), ("TEXTCOLOR", (0, 1), (-1, -1), colors.HexColor("#E8D48A")), ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#2A2200"))]))
    elements.append(tool_table)
    elements.append(Spacer(1, 6 * mm))

    elements.append(Paragraph("E. Alerts", h2))
    alert_stats_rows = [["Metric", "Value"]] + [[k, str(v)] for k, v in report["alerts_stats"].items()]
    ast = Table(alert_stats_rows, colWidths=[80 * mm, 85 * mm])
    ast.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FFD700")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.black), ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#13100A")), ("TEXTCOLOR", (0, 1), (-1, -1), colors.HexColor("#E8D48A")), ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#2A2200"))]))
    elements.append(ast)
    elements.append(Spacer(1, 6 * mm))

    elements.append(Paragraph("F. Settings Snapshot", h2))
    setting_rows = [["Setting", "Value"]] + [[r["Setting"], str(r["Value"])] for r in report["settings_snapshot"]]
    sst = Table(setting_rows, colWidths=[80 * mm, 85 * mm])
    sst.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FFD700")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.black), ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#13100A")), ("TEXTCOLOR", (0, 1), (-1, -1), colors.HexColor("#E8D48A")), ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#2A2200"))]))
    elements.append(sst)
    elements.append(Spacer(1, 6 * mm))

    elements.append(Paragraph("G. Recommendations", h2))
    for section_name in ["high", "medium", "gdpr", "device"]:
        for rec in report["recommendations"][section_name]:
            elements.append(Paragraph(f"<b>{rec['priority']}</b> | {rec['target']} | {rec['text']}", body))
            elements.append(Spacer(1, 2 * mm))
    for text in report["recommendations"]["systemic"]:
        elements.append(Paragraph(f"- {text}", body))

    def decorate(canvas, _doc):
        canvas.saveState()
        canvas.setFillColor(colors.HexColor("#0C0A04"))
        canvas.rect(0, 0, A4[0], A4[1], stroke=0, fill=1)
        canvas.setFillColor(colors.HexColor("#E8D48A"))
        page = canvas.getPageNumber()
        canvas.setFont("Helvetica", 9)
        canvas.drawString(12 * mm, A4[1] - 10 * mm, report["org_name"])
        canvas.drawCentredString(A4[0] / 2, A4[1] - 10 * mm, "Shadow IT DNA Map")
        canvas.drawRightString(A4[0] - 12 * mm, A4[1] - 10 * mm, f"Page {page}")
        canvas.drawString(12 * mm, 8 * mm, "CONFIDENTIAL — Internal Use Only")
        canvas.drawRightString(A4[0] - 12 * mm, 8 * mm, report["generated_at"].strftime("%Y-%m-%d %H:%M:%S"))
        canvas.restoreState()

    doc.build(elements, onFirstPage=decorate, onLaterPages=decorate)
    buff.seek(0)
    return buff.getvalue()


def render_sidebar():
    with st.sidebar:
        st.markdown(f"<h2 style='color:{THEME['color']}'>🔍 Shadow IT DNA Map</h2>", unsafe_allow_html=True)
        st.markdown(
            f"<p style='color:{THEME['muted']}; margin-top:-8px;'>02 — IT Management</p>",
            unsafe_allow_html=True,
        )
        st.markdown("<hr>", unsafe_allow_html=True)
        st.page_link("app.py", label="Home")
        st.page_link("pages/01_dashboard.py", label="Live Exposure Dashboard")
        st.page_link("pages/02_upload.py", label="DNS Log Upload & Parsing")
        st.page_link("pages/03_tools.py", label="Shadow Tool Browser")
        st.page_link("pages/04_alerts.py", label="Alerts Manager")
        st.page_link("pages/05_settings.py", label="Settings & Reference")
        st.page_link("pages/06_report.py", label="📄 Master Report")


def render_preview(report):
    st.subheader("SECTION A — Executive Summary")
    exec_df = pd.DataFrame(
        [
            ["Organisation name", report["executive"]["organisation"]],
            ["Report generated date/time", report["executive"]["generated"]],
            ["Report period", report["executive"]["period"]],
            ["Total sessions analysed", report["executive"]["total_sessions"]],
            ["Total DNS queries analysed", report["executive"]["total_queries"]],
            ["Total unique domains seen", report["executive"]["unique_domains"]],
            ["Total unique source IPs seen", report["executive"]["unique_source_ips"]],
            ["Total shadow tools ever detected", report["executive"]["total_shadow_tools"]],
            ["Overall Exposure Score", report["executive"]["overall_score"]],
            ["Overall Risk Verdict", report["executive"]["overall_verdict"]],
            ["Total Cumulative App Usage Time", report["executive"]["total_cumulative_app_usage_time"]],
            ["Average Session Duration", report["executive"]["average_session_duration"]],
            ["Longest Working Session", report["executive"]["longest_working_session"]],
            ["Most Active Hour of Day", report["executive"]["most_active_hour_of_day"]],
            ["Total Pages Visited (all sessions)", report["executive"]["total_pages_visited_all_sessions"]],
            ["Most Visited Page", report["executive"]["most_visited_page"]],
            ["Total Shadow App Usage Time (all tools, all sessions)", report["executive"]["total_shadow_app_usage_time"]],
            ["Average Shadow App Usage Per Session", report["executive"]["average_shadow_app_usage_per_session"]],
            ["Single Most Used Shadow App", report["executive"]["single_most_used_shadow_app"]],
            ["Longest Uninterrupted Shadow Tool Session", report["executive"]["longest_uninterrupted_shadow_tool_session"]],
            ["Peak Shadow Activity Hour", report["executive"]["peak_shadow_activity_hour"]],
            ["% of Workday Spent on Shadow Tools (est.)", report["executive"]["pct_workday_spent_on_shadow_tools"]],
            ["After-Hours Shadow Usage Detected", report["executive"]["after_hours_shadow_usage_detected"]],
        ],
        columns=["Field", "Value"],
    )
    st.dataframe(exec_df, use_container_width=True)
    trend = report["executive"]["trend"]
    st.write(f"Score Trend: {trend}")
    st.write(
        f"Highest Session: {report['executive']['highest']['parsed_at'] if report['executive']['highest'] else 'N/A'} / {report['executive']['highest']['score'] if report['executive']['highest'] else 'N/A'}"
    )
    st.write(
        f"Lowest Session: {report['executive']['lowest']['parsed_at'] if report['executive']['lowest'] else 'N/A'} / {report['executive']['lowest']['score'] if report['executive']['lowest'] else 'N/A'}"
    )

    st.subheader("SECTION B — Per-Session Breakdown")
    for idx, s in enumerate(report["sessions"], start=1):
        with st.expander(f"B.{idx} — Session [{idx}]: {s['parsed_at']}", expanded=(idx == 1)):
            ss_df = pd.DataFrame(
                [
                    ["Session ID", s["session_id"]],
                    ["Parsed At", s["parsed_at"]],
                    ["Source File", s["source_file"]],
                    ["File Size", s["file_size"]],
                    ["Total DNS Queries", s["total_queries"]],
                    ["Unique Domains", s["unique_domains"]],
                    ["Unique Source IPs", s["unique_source_ips"]],
                    ["Shadow Tools Detected", s["shadow_tools"]],
                    ["HIGH Risk Tools", s["high_count"]],
                    ["MEDIUM Risk Tools", s["medium_count"]],
                    ["LOW Risk Tools", s["low_count"]],
                    ["GDPR-flagged Tools", s["gdpr_count"]],
                    ["Exposure Score", f"{s['score']} ({s['score_label']})"],
                    ["Auto-alerts Created", s["auto_alerts"]],
                    ["App Opened At", dt_text(s.get("app_opened_at"))],
                    ["Analysis Started At", dt_text(s.get("analysis_started_at"))],
                    ["Analysis Ended At", dt_text(s.get("analysis_ended_at"))],
                    ["App Closed At", dt_text(s.get("app_closed_at"))],
                    ["Analysis Duration", s.get("analysis_duration_fmt", "N/A")],
                    ["Total App Session Duration", s.get("total_app_duration_fmt", "N/A")],
                    ["Pages Visited", pages_text(s.get("pages_visited", []))],
                    ["Page Visit Count", s.get("page_visit_count", 0)],
                ],
                columns=["Field", "Value"],
            )
            st.dataframe(ss_df, use_container_width=True)

            tools_df = pd.DataFrame(
                [
                    {
                        "#": i + 1,
                        "Tool Name": t["tool_name"],
                        "Category": t["category"],
                        "Risk": t["risk"],
                        "Query Count": t["query_count"],
                        "First Seen": dt_text((t.get("duration") or {}).get("first_query_at")),
                        "Last Seen": dt_text((t.get("duration") or {}).get("last_query_at")),
                        "Total Span": (t.get("duration") or {}).get("total_span_fmt", "00h 00m 00s"),
                        "Active Minutes": int((t.get("duration") or {}).get("active_minutes", 0) or 0),
                        "Queries/Min": float((t.get("duration") or {}).get("queries_per_minute", 0) or 0),
                        "Peak Hour": (
                            f"{int((t.get('duration') or {}).get('peak_usage_hour', 0)):02d}:00-{(int((t.get('duration') or {}).get('peak_usage_hour', 0)) + 1) % 24:02d}:00"
                            if (t.get("duration") or {}).get("peak_usage_hour") is not None
                            else "N/A"
                        ),
                        "% of Total Queries": round((t["query_count"] / max(1, s["total_queries"])) * 100, 2),
                        "Unique IPs": t["unique_ips"],
                        "GDPR": "YES" if t["gdpr"] else "No",
                        "Risk Reasons": "; ".join(t["risk_reasons"]) or "N/A",
                        "Recommended Alternative": t["alternative"],
                    }
                    for i, t in enumerate(s["tools"])
                ]
            )
            st.markdown("Shadow Tools Detected in This Session")
            st.dataframe(tools_df, use_container_width=True)

            if s["tools"]:
                most_used = max(s["tools"], key=lambda x: (x.get("duration") or {}).get("total_span_seconds", 0))
                most_queried = max(s["tools"], key=lambda x: x.get("query_count", 0))
                longest_active = max(s["tools"], key=lambda x: (x.get("duration") or {}).get("active_minutes", 0))
                peak_counter = Counter(
                    (t.get("duration") or {}).get("peak_usage_hour")
                    for t in s["tools"]
                    if (t.get("duration") or {}).get("peak_usage_hour") is not None
                )
                peak_hour = peak_counter.most_common(1)[0][0] if peak_counter else None
                st.markdown("Duration Summary")
                st.write(
                    f"Most Used App (by duration): {most_used['tool_name']} - {(most_used.get('duration') or {}).get('total_span_fmt', '00h 00m 00s')}"
                )
                st.write(f"Most Queried App (by count): {most_queried['tool_name']} - {most_queried['query_count']:,} queries")
                st.write(
                    f"Longest Active App: {longest_active['tool_name']} - {int((longest_active.get('duration') or {}).get('active_minutes', 0) or 0)} active minutes"
                )
                st.write(
                    f"Peak Shadow Usage Hour: {peak_hour:02d}:00-{(peak_hour + 1) % 24:02d}:00"
                    if peak_hour is not None
                    else "Peak Shadow Usage Hour: N/A"
                )

            st.markdown("Top 3 Most Queried Domains")
            if s["top_domains"]:
                st.table(pd.DataFrame(s["top_domains"], columns=["Domain", "Query Count"]))
            else:
                st.write("N/A")

            st.markdown("Source IPs Active in This Session")
            ip_df = pd.DataFrame(
                [
                    {
                        "IP Address": ip["ip"],
                        "Query Count": ip["query_count"],
                        "Tools Accessed": ", ".join(ip["tools_accessed"]) or "N/A",
                        "Risk Level of Tools": ip["risk_level"],
                    }
                    for ip in s["ip_rows"]
                ]
            )
            st.dataframe(ip_df, use_container_width=True)

    st.subheader("SECTION C — All Users / Devices (Cross-Session)")
    st.markdown("C1. Device Risk Roster")
    c_df = pd.DataFrame(report["devices"])
    if not c_df.empty and "repeat_offender" in c_df.columns:
        c_df = c_df.drop(columns=["repeat_offender"])
    st.dataframe(c_df, use_container_width=True)

    st.markdown("C2. Power Users (Top 10)")
    st.dataframe(pd.DataFrame(report["power_users"]), use_container_width=True)
    st.markdown("C3. GDPR Risk Devices")
    st.dataframe(pd.DataFrame(report["gdpr_devices"]), use_container_width=True)
    st.markdown("C4. Repeat Offenders")
    st.dataframe(pd.DataFrame(report["repeat_offenders"]), use_container_width=True)

    st.markdown("C5. User Working Hours Analysis")
    c5_df = pd.DataFrame(report["user_session_timing"]["rows"])
    if not c5_df.empty:
        c5_view = c5_df.drop(columns=["Analysis Duration Sec", "Total App Duration Sec"], errors="ignore")
        st.dataframe(c5_view, use_container_width=True)
    else:
        st.info("No user timing records available.")
    c5_summary = report["user_session_timing"]["summary"]
    st.table(
        pd.DataFrame(
            [
                ["Total cumulative time", c5_summary["total_cumulative_time_fmt"]],
                ["Average session duration", c5_summary["average_session_duration_fmt"]],
                [
                    "Longest session",
                    (
                        f"{dt_text(c5_summary['longest_session']['parsed_at'])} - {format_duration(int(c5_summary['longest_session'].get('total_app_duration_sec', 0) or 0))}"
                        if c5_summary.get("longest_session")
                        else "N/A"
                    ),
                ],
                [
                    "Shortest session",
                    (
                        f"{dt_text(c5_summary['shortest_session']['parsed_at'])} - {format_duration(int(c5_summary['shortest_session'].get('total_app_duration_sec', 0) or 0))}"
                        if c5_summary.get("shortest_session")
                        else "N/A"
                    ),
                ],
                ["Most visited page", f"{c5_summary['most_visited_page']} ({c5_summary['most_visited_page_count']})"],
                ["Peak usage hour", c5_summary["peak_usage_hour"]],
            ],
            columns=["Metric", "Value"],
        )
    )

    st.markdown("C6. Per-Device Per-App Duration Table")
    c6_df = pd.DataFrame(report["per_device_app_duration"])
    if not c6_df.empty:
        st.dataframe(c6_df.drop(columns=["Total Duration Sec"], errors="ignore"), use_container_width=True)
    else:
        st.info("No per-device per-app duration data available.")

    st.markdown("C7. Device Usage Heatmap Data")
    c7_df = pd.DataFrame(report["device_usage_heatmap"])
    if not c7_df.empty:
        st.dataframe(c7_df, use_container_width=True)
    else:
        st.info("No hourly heatmap data available.")

    st.subheader("SECTION D — Shadow Tool Intelligence (Master Catalog)")
    st.markdown("D1. All Detected Tools — Master Table")
    master_df = pd.DataFrame([{k: v for k, v in t.items() if not k.startswith("_")} for t in report["master_tools"]])
    st.dataframe(master_df, use_container_width=True)
    st.markdown("D2. Tool Frequency Analysis")
    st.write(f"Most persistent tool: {report['most_persistent']['Tool Name'] if report['most_persistent'] else 'N/A'}")
    st.write(
        f"Fastest growing tool: {report['fastest_growing']['Tool Name'] if report['fastest_growing'] else 'N/A'} (Delta {report['fastest_growth_delta']})"
    )
    st.write(f"New tools in latest session: {', '.join(report['new_tools_latest']) or 'None'}")
    st.write(f"Tools disappeared in latest session: {', '.join(report['disappeared_tools']) or 'None'}")
    st.markdown("D3. Category Breakdown")
    st.dataframe(pd.DataFrame(report["category_rows"]), use_container_width=True)
    st.markdown("D4. GDPR Exposure Summary")
    st.dataframe(pd.DataFrame(report["gdpr_tool_rows"]), use_container_width=True)
    st.write(f"Total personal data exposure estimate: {report['personal_data_exposure_est']}")

    st.markdown("D5. App Duration Intelligence — Master Table")
    d5_df = pd.DataFrame(report["duration_intelligence"]["master_table"])
    if not d5_df.empty:
        st.dataframe(
            d5_df[
                [
                    "Tool Name",
                    "Category",
                    "Risk Level",
                    "Total Duration (all sessions)",
                    "Avg Duration Per Session",
                    "Longest Single Session Duration",
                    "Total Active Minutes",
                    "Total Queries (all sessions)",
                    "Avg Queries/Min",
                    "Peak Usage Hour",
                    "Sessions Appeared In",
                    "After Hours %",
                    "After Hours Flag",
                ]
            ],
            use_container_width=True,
        )
    st.markdown("D6. Duration Rankings")
    ranks = report["duration_intelligence"]["rankings"]
    if ranks["top_duration"]:
        for idx, row in enumerate(ranks["top_duration"][:3], start=1):
            medal = "🥇" if idx == 1 else "🥈" if idx == 2 else "🥉"
            st.write(f"{medal} {row['Tool Name']} - {row['Total Duration (all sessions)']} across {row['Sessions Appeared In']} sessions")
    if ranks["top_intensive"]:
        top_int = ranks["top_intensive"][0]
        st.write(f"⚡ Most Intensive (queries/min): {top_int['Tool Name']} - {top_int['Avg Queries/Min']} queries/min")
    if report["executive"].get("longest_uninterrupted_shadow_tool_session"):
        st.write(f"🕐 Longest Single Sitting: {report['executive']['longest_uninterrupted_shadow_tool_session']}")
    if ranks["top_after_hours"]:
        top_after = ranks["top_after_hours"][0]
        st.write(f"🌙 Most After-Hours Usage: {top_after['Tool Name']} - {top_after['After Hours %']}% after hours")

    st.markdown("D7. App Usage Timeline")
    timeline_rows = []
    for item in report["duration_intelligence"]["timeline"]:
        for point in item.get("points", []):
            timeline_rows.append(
                {
                    "Tool": item.get("tool"),
                    "Session Date": dt_text(point[0]),
                    "Total Duration Seconds": int(point[1]),
                }
            )
    st.dataframe(pd.DataFrame(timeline_rows), use_container_width=True)

    st.subheader("SECTION E — Alerts — Full History")
    st.table(pd.DataFrame(list(report["alerts_stats"].items()), columns=["Metric", "Value"]))
    st.dataframe(pd.DataFrame(report["alerts"]), use_container_width=True)
    st.markdown("Unresolved HIGH Alerts")
    st.dataframe(pd.DataFrame(report["unresolved_high"]), use_container_width=True)

    st.subheader("SECTION F — Settings Snapshot")
    st.table(pd.DataFrame(report["settings_snapshot"]))

    st.subheader("SECTION G — Recommendations")
    st.markdown("G1. Immediate Actions (HIGH risk tools)")
    for r in report["recommendations"]["high"]:
        st.write(f"- 🔴 {r['target']} — {r['text']}")
    st.markdown("G2. Review Required (MEDIUM risk tools)")
    for r in report["recommendations"]["medium"]:
        st.write(f"- 🟡 {r['target']} — {r['text']}")
    st.markdown("G3. GDPR Compliance Actions")
    for r in report["recommendations"]["gdpr"]:
        st.write(f"- ⚠️ {r['target']} — {r['text']}")
    st.markdown("G4. Device-Level Actions")
    for r in report["recommendations"]["device"]:
        st.write(f"- 📍 Device {r['target']} — {r['text']}")
    st.markdown("G5. Systemic Recommendations")
    for r in report["recommendations"]["systemic"]:
        st.write(f"- {r}")


st.set_page_config(page_title="Master Report", page_icon="📄", layout="wide")
apply_theme(st)
render_sidebar()

st.title("📄 Master Report — Full System Export")
st.caption(
    "Complete shadow IT intelligence report covering all sessions, tools, alerts, devices, and recommendations."
)

warnings = []
sessions_json = safe_json(SESSIONS_PATH, {"sessions": []}, "Sessions section", warnings)
alerts_json = safe_json(ALERTS_PATH, {"alerts": []}, "Alerts section", warnings)
settings_json = safe_json(SETTINGS_PATH, {}, "Settings section", warnings)
catalog_json = safe_json(CATALOG_PATH, {"tools": []}, "Catalog section", warnings)

for w in warnings:
    st.warning(w)

all_sessions = sessions_json.get("sessions", []) if isinstance(sessions_json, dict) else []
alerts = alerts_json.get("alerts", []) if isinstance(alerts_json, dict) else []
settings = settings_json if isinstance(settings_json, dict) else {}
catalog_tools = catalog_json.get("tools", []) if isinstance(catalog_json, dict) else []

if not all_sessions:
    st.info("No sessions found yet. Upload or generate DNS logs first to build the master report.")
    if st.button("Go to Upload Page"):
        st.switch_page("pages/02_upload.py")
    st.stop()

session_options = {
    f"{idx + 1}. {s.get('name', 'Session')} | {s.get('uploaded_at', 'N/A')}": idx
    for idx, s in enumerate(all_sessions)
}

scope = st.radio("Report Scope", ["All Sessions", "Single Session"], horizontal=True)
selected_idx = None
if scope == "Single Session":
    selected_label = st.selectbox("Single Session", list(session_options.keys()))
    selected_idx = session_options[selected_label]

default_org = settings.get("organisation_name", settings.get("org_name", ""))
org_name = st.text_input("Organisation Name", value=default_org)
report_title = st.text_input("Report Title", value="Shadow IT DNA Map — Full System Report")

if "master_report_ready" not in st.session_state:
    st.session_state["master_report_ready"] = False

if st.button("Generate Report Preview", type="primary"):
    st.session_state["master_report_ready"] = True

if st.session_state["master_report_ready"]:
    selected_sessions = all_sessions if scope == "All Sessions" else [all_sessions[selected_idx]]
    report = compute_report(
        selected_sessions=selected_sessions,
        all_sessions=all_sessions,
        alerts=alerts,
        settings=settings,
        catalog_tools=catalog_tools,
        org_name=org_name or "UnknownOrg",
        report_title=report_title,
    )

    render_preview(report)

    safe_org = "".join(ch for ch in (org_name or "UnknownOrg") if ch.isalnum() or ch in ("-", "_")) or "UnknownOrg"
    date_str = datetime.now().strftime("%Y%m%d")
    excel_name = f"shadow_it_MASTER_REPORT_{safe_org}_{date_str}.xlsx"
    pdf_name = f"shadow_it_MASTER_REPORT_{safe_org}_{date_str}.pdf"

    c1, c2 = st.columns(2)
    with c1:
        try:
            xlsx_bytes = build_excel(report)
            st.download_button("Download Excel (.xlsx)", data=xlsx_bytes, file_name=excel_name, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
        except Exception as ex:  # noqa: BLE001
            st.error(f"Excel generation failed: {ex}")
    with c2:
        try:
            pdf_bytes = build_pdf(report)
            st.download_button("Download PDF (.pdf)", data=pdf_bytes, file_name=pdf_name, mime="application/pdf", use_container_width=True)
        except Exception as ex:  # noqa: BLE001
            st.error(f"PDF generation failed: {ex}")
